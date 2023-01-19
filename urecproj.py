import xlsxwriter
import openpyxl as op
import os

################################################################################
# Create template

def template(workbook, title):

    # Default formatting
    workbook.formats[0].set_font_size(10)
    workbook.formats[0].set_align('center')
    workbook.formats[0].set_align('vcenter')

    # Add formatting
    worksheet = workbook.add_worksheet()
    worksheet.set_column(1,16,15)
    worksheet.set_column(7,7,2)
    worksheet.set_column(12,12,2)

    title_format = workbook.add_format()
    title_format.set_font_size(22)
    title_format.set_font_name('Elephant Pro')
    title_format.set_align('center')
    title_format.set_align('vcenter')

    subtitle_format = workbook.add_format()
    subtitle_format.set_font_color('white')
    subtitle_format.set_bg_color('black')
    subtitle_format.set_align('center')
    subtitle_format.set_align('vcenter')
    subtitle_format.set_bold()
    subtitle_format.set_border(1)

    labels_format = workbook.add_format()
    labels_format.set_align('center')
    labels_format.set_align('vcenter')
    labels_format.set_bold()
    labels_format.set_border(1)

    black_format = workbook.add_format()
    black_format.set_bg_color('black')
    black_format.set_border(1)

    worksheet.merge_range('B2:L3', 'Merged')
    worksheet.merge_range('B5:L5', 'Merged')
    worksheet.merge_range('N5:Q5', 'Merged')

    worksheet.write('B7','',black_format)
    worksheet.write('I7','',black_format)
    worksheet.write('K7','',black_format)

    # Arrays for default values
    weekdays = ('Monday','Tuesday','Wednesday','Thurday','Friday')
    students = ('Name','Return date','Hours desired', 'Notes')

    # Fill text
    worksheet.write('B2',title,title_format)
    worksheet.write('B5','AVAILABILITIES',subtitle_format)
    worksheet.write('N5','STUDENTS',subtitle_format)

    row = 7
    col = 3
    for value in weekdays:
        worksheet.write(row-1,col-1,value,labels_format) # -1 because row/col values start from 0
        col+=1

    worksheet.write('J7','Saturday',labels_format)
    worksheet.write('L7', 'Sunday',labels_format)

    row = 7
    col = 14
    for value in students:
        worksheet.write(row-1,col-1,value,labels_format)
        col+=1

    return worksheet

################################################################################
# Read shifts

def shifts(sh, nm, shift_li):

    # create list for mon~sun shifts
    for i in range(7):
        shift_li.append([])
        if i<=4:
            for j in range(6):
                shift_li[i].append([[],[],[]])
        if i>=5:
            for j in range(4):
                shift_li[i].append([[],[],[]])
    # shift_li:
    #    1. column (7)
    #    2. row (6 for col 0~4(weekday), 4 for col 5~6(weekend))
    #    3. [green][orange][red]

    # weekday
    for c in range(5):
        for r in range(6):
            bgcol = sh.cell(row=7+r, column=2+c).fill.start_color.index
            if bgcol=='FF00B050':
                shift_li[c][r][0].append(nm) # Green
            if bgcol=='FFF79645':
                shift_li[c][r][1].append(nm) # Orange
            if bgcol=='FFFF0000':
                shift_li[c][r][2].append(nm) # Red

    # saturday
    for r in range(4):
        bgcol = sh.cell(row=15+r, column=2).fill.start_color.index
        if bgcol=='FF00B050':
            shift_li[5][r][0].append(nm) # Green
        if bgcol=='FFF79645':
            shift_li[5][r][1].append(nm) # Orange
        if bgcol=='FFFF0000':
            shift_li[5][r][2].append(nm) # Red

    # sunday
    for r in range(4):
        bgcol = sh.cell(row=15+r, column=4).fill.start_color.index
        if bgcol=='FF00B050':
            shift_li[6][r][0].append(nm) # Green
        if bgcol=='FFF79645':
            shift_li[6][r][1].append(nm) # Orange
        if bgcol=='FFFF0000':
            shift_li[6][r][2].append(nm) # Red

    # weekday: B7~F12
    # saturday: B15~B18
    # sunday: D15~D18
    # SPN: J7~N9

    return shift_li

################################################################################
# Read excel files (starting with student information)

def read(path):

    names_li = [] # Empty list for names
    returns_li = []
    hours_li = []
    notes_li = []

    files = os.listdir(path)
    shift_li = []
    for f in files:
        if f.endswith('.xlsx'):
            print("Reading file "+f+"...")
            wb = op.load_workbook(path+"/"+f)
            sh = wb.active

            # Read student information
            names_li.append(str(sh["c2"].value))
            returns_li.append(str(sh["c3"].value))
            hours_li.append(str(sh["c4"].value))
            notes_li.append(str(sh["e15"].value))

            nm = str(sh["c2"].value)
            # Read shift information
            shift_li = shifts(sh, nm, shift_li) # return this along with the students list

        else:
            print("Skipping file "+f+"...")

    students_li = [names_li, returns_li, hours_li, notes_li]
    final_li = [students_li, shift_li]

    return final_li

################################################################################
# Read excel files

def fill(worksheet, info):
    # info = final_li

    left_format = workbook.add_format()
    left_format.set_align('left')
    green = workbook.add_format()
    green.set_font_color('green')
    orange = workbook.add_format()
    orange.set_font_color('orange')
    red = workbook.add_format()
    red.set_font_color('red')

    # fill student information
    row = 8
    col = 14
    for type in info[0]: # student_li
        for value in type:
            worksheet.write(row-1, col-1, value, left_format)
            row+=1
        col+=1
        row=8

    # fill shift information
    row = 8
    col = 3
    weekday_breaks = ['7a-10a', '10a-1p', '1p-4p', '4p-7p', '7p-10p', '10p-1a']
    sat_breaks = ['8a-12a', '12p-3p', '3p--6p', '6p-10p']
    sun_breaks = ['12p-4p', '4p-7p', '7p-10p', '10p-1a']

    for c in info[1]: # columns in shift_li
        for r in c:
            # weekday
            if info[1].index(c)<=4:
                if col == 3:
                    worksheet.write(row-1, col-2,weekday_breaks[c.index(r)])
                for nms in r[0]:
                    worksheet.write(row-1, col-1, nms, green)
                    row+=1
                for nms in r[1]:
                    worksheet.write(row-1, col-1, nms, orange)
                    row+=1
                for nms in r[2]:
                    worksheet.write(row-1, col-1, nms, red)
                    row+=1
            # saturday
            if info[1].index(c)==5:
                for nms in r[0]:
                    worksheet.write(row-1, 9, nms, green)
                    row+=1
                for nms in r[1]:
                    worksheet.write(row-1, 9, nms, orange)
                    row+=1
                for nms in r[2]:
                    worksheet.write(row-1, 9, nms, red)
                    row+=1
            # sunday
            if info[1].index(c)==6:
                for nms in r[0]:
                    worksheet.write(row-1, 11, nms, green)
                    row+=1
                for nms in r[1]:
                    worksheet.write(row-1, 11, nms, orange)
                    row+=1
                for nms in r[2]:
                    worksheet.write(row-1, 11, nms, red)
                    row+=1
            row+=1
        col+=1
        row=8

################################################################################
# Start program

print('\n#####################')
print("UREC Scheduler\nUpdated: -\nWritten by Zia Kim")
print('#####################\n')
inp = input("Enter semester and year (ex. Spring 2023): ")
title = inp + " Availability"

#Create excel file with default template
workbook = xlsxwriter.Workbook(title+".xlsx")
worksheet = template(workbook, title)

# Read info
path = input("Enter path: ")
path = "/Users/zia/Downloads/test"
info = read(path)

# Fill info
fill(worksheet, info)

workbook.close()